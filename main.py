import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Создаем директорию для изображений, если она не существует
if not os.path.exists('images'):
    os.makedirs('images')

# Подключение к SQLite (если базы данных нет, она будет создана)
conn = sqlite3.connect('your_database.db')
cursor = conn.cursor()

# Создание таблицы, если она не существует
cursor.execute('''
CREATE TABLE IF NOT EXISTS your_table (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    order_number TEXT,
    client_name TEXT,
    order_date TEXT,
    image_path TEXT
)
''')

# Чтение данных из Excel файла
wb = load_workbook('1.xlsx')
sheet = wb.active

# Словарь для хранения изображений по номеру заказа
images_dict = {img.anchor._from.row: img for img in sheet._images}

for row in sheet.iter_rows(min_row=2, max_row=11, values_only=False):  # Проходим по строкам с 2 по 11
    order_number = row[0].value  # Номер заказа (столбец A)
    client_name = row[5].value    # Имя клиента (столбец F)
    order_date = row[7].value     # Дата заказа (столбец H)

    # Вставка данных в таблицу
    cursor.execute('''
    INSERT INTO your_table (order_number, client_name, order_date, image_path)
    VALUES (?, ?, ?, ?)
    ''', (order_number, client_name, order_date, None))  # Временно устанавливаем image_path в None

    # Получаем ID последней вставленной записи
    record_id = cursor.lastrowid

    # Если в словаре есть изображение для текущей строки, сохраняем его
    if row[6].row in images_dict:  # Проверяем, есть ли изображение для текущей строки
        image = images_dict[row[6].row]
        image_path = f'images/{order_number}.png'  # Имя файла соответствует номеру заказа
        with open(image_path, 'wb') as img_file:
            img_file.write(image._data())  # Используем метод _data() для получения байтов изображения
        # Обновляем запись с путем к изображению
        cursor.execute('''
        UPDATE your_table
        SET image_path = ?
        WHERE id = ?
        ''', (image_path, record_id))

# Сохраняем изменения и закрываем соединение
conn.commit()
cursor.close()
conn.close()
